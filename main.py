import pandas as pd
from datetime import datetime
from tkinter import filedialog
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from database import Database
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class MainWindow:
    def delete_stock_item(self):
        sel = self.tree_stock.selection()
        if not sel:
            messagebox.showwarning("Внимание", "Выберите запись для удаления")
            return

        values = self.tree_stock.item(sel[0])['values']
        stock_id = values[0]
        med_name = values[2]
        qty = values[3]

        if not messagebox.askyesno("Подтверждение удаления",
                                  f"Вы действительно хотите удалить:\n\n"
                                  f"{med_name}\n"
                                  f"Количество: {qty}\n\n"
                                  f"Это действие нельзя отменить!"):
            return

        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM stock WHERE id = ?", (stock_id,))
            conn.commit()
            conn.close()

            messagebox.showinfo("Успешно", "Запись удалена")
            self.load_stock()   # обновляем таблицу
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось удалить:\n{str(e)}")

    def __init__(self):
        self.db = Database()
        self.db.add_test_data()

        self.root = tk.Tk()
        self.root.title("Информационная система МСЧ ФСИН")
        self.root.geometry("1320x780")
        self.root.minsize(1200, 700)

        title = tk.Label(self.root, text="Информационная система медико-санитарной части ФСИН",
                         font=("Arial", 16, "bold"))
        title.pack(pady=10)

        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=5)

        self.create_tabs()

        self.status = tk.Label(self.root, text="Готов к работе", relief=tk.SUNKEN, anchor=tk.W)
        self.status.pack(side=tk.BOTTOM, fill=tk.X)

    def create_tabs(self):
        self.tab_stock = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_stock, text="📦 Склад и остатки")
        self.create_stock_tab()

        self.tab_movements = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_movements, text="🔄 Перемещения")
        self.create_movements_tab()

        self.tab_expenditures = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_expenditures, text="📉 Расходование")
        self.create_expenditures_tab()

        self.tab_prisoners = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_prisoners, text="🧍 Учёт медицинской помощи")
        self.create_prisoners_tab()

        self.tab_reports = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_reports, text="📊 Отчёты")
        self.create_reports_tab()   # ← Теперь вызов корректный

    # ====================== ОТЧЁТЫ ======================
    def create_reports_tab(self):
        top = ttk.Frame(self.tab_reports)
        top.pack(fill="x", padx=15, pady=20)

        # Центральная кнопка
        btn_frame = ttk.Frame(top)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="📊 Сформировать отчёт",
                   command=self.open_report_dialog,
                   width=30).pack()

        # Инструкция
        ttk.Label(self.tab_reports,
                  text="Нажмите кнопку выше и выберите параметры отчёта\n"
                       "Отчёт будет сохранён в удобном Excel формате.",
                  font=("Arial", 10), foreground="gray", justify="center").pack(pady=60)

        # Валидация дат для окна отчётов
        self.validate_date_cmd = (self.root.register(self.validate_date), '%P')

    def load_report_filters(self):
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM med_parts ORDER BY name")
        parts = ["Все медчасти"] + [r["name"] for r in cursor.fetchall()]
        self.report_part_combo['values'] = parts
        self.report_part_combo.current(0)
        conn.close()

    def generate_report(self):
        pass  # теперь используется open_report_dialog

    def open_report_dialog(self):
        ReportDialog(self.root, self)

    def report_expenditures(self, date_from, date_to, part_filter):
        conn = self.db.get_connection()

        # Основной детальный отчёт
        query = '''
                SELECT
                    DATE (e.date) as "Дата", mp.name as "Медчасть", m.name as "Препарат", COALESCE (p.full_name, '—') as "Заключённый", e.quantity as "Кол-во", e.reason as "Причина", e.comment as "Комментарий"
                FROM expenditures e
                    JOIN stock s \
                ON e.stock_id = s.id
                    JOIN med_parts mp ON s.med_part_id = mp.id
                    JOIN medications m ON s.medication_id = m.id
                    LEFT JOIN prisoners p ON e.prisoner_id = p.id
                WHERE DATE (e.date) BETWEEN ? AND ? \
                '''
        params = [date_from, date_to]
        if part_filter != "Все медчасти":
            query += " AND mp.name = ?"
            params.append(part_filter)

        df_detail = pd.read_sql_query(query, conn, params=params)

        # === Сводка по заключённым ===
        if part_filter != "Все медчасти":
            summary_query = '''
                            SELECT COALESCE(p.full_name, '—')    as "Заключённый", \
                                   SUM(e.quantity)               as "Всего израсходовано", \
                                   COUNT(*)                      as "Кол-во записей", \
                                   GROUP_CONCAT(DISTINCT m.name) as "Препараты"
                            FROM expenditures e
                                     JOIN stock s ON e.stock_id = s.id
                                     JOIN med_parts mp ON s.med_part_id = mp.id
                                     JOIN medications m ON s.medication_id = m.id
                                     LEFT JOIN prisoners p ON e.prisoner_id = p.id
                            WHERE DATE (e.date) BETWEEN ? AND ? AND mp.name = ?
                            GROUP BY p.full_name
                            ORDER BY "Всего израсходовано" DESC \
                            '''
            df_summary = pd.read_sql_query(summary_query, conn, params=[date_from, date_to, part_filter])
        else:
            summary_query = '''
                            SELECT COALESCE(p.full_name, '—')    as "Заключённый", \
                                   SUM(e.quantity)               as "Всего израсходовано", \
                                   COUNT(*)                      as "Кол-во записей", \
                                   GROUP_CONCAT(DISTINCT m.name) as "Препараты"
                            FROM expenditures e
                                     JOIN stock s ON e.stock_id = s.id
                                     JOIN medications m ON s.medication_id = m.id
                                     LEFT JOIN prisoners p ON e.prisoner_id = p.id
                            WHERE DATE (e.date) BETWEEN ? AND ?
                            GROUP BY p.full_name
                            ORDER BY "Всего израсходовано" DESC \
                            '''
            df_summary = pd.read_sql_query(summary_query, conn, params=[date_from, date_to])

        conn.close()

        if df_detail.empty:
            messagebox.showinfo("Информация", "Нет данных за выбранный период")
            return

        # === Сохранение с красивым оформлением ===
        default_filename = f"Расходование_медикаментов_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=default_filename,
            filetypes=[("Excel файлы", "*.xlsx")]
        )
        if not file_path:
            return

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Детальный лист
            df_detail.to_excel(writer, sheet_name="Детальный расход", index=False)
            # Сводный лист
            if not df_summary.empty:
                df_summary.to_excel(writer, sheet_name="Сводка по заключённым", index=False)

            # Применяем красивое форматирование к обоим листам
            for sheet_name in writer.sheets:
                ws = writer.sheets[sheet_name]
                df_current = df_detail if sheet_name == "Детальный расход" else df_summary

                # Автоширина колонок
                for idx, col in enumerate(df_current.columns):
                    series = df_current[col].astype(str)
                    max_length = max(series.str.len().max() if not series.empty else 0, len(col)) + 3
                    ws.column_dimensions[chr(65 + idx)].width = min(max_length, 60)

                # Красивые заголовки
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                alignment = Alignment(horizontal="center", vertical="center")

                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = alignment

                # Границы ячеек
                thin = Side(style='thin')
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                for row in ws.iter_rows():
                    for cell in row:
                        cell.border = border

        messagebox.showinfo("✅ Успешно", f"Отчёт успешно сохранён!\n\n{file_path}\n(2 листа)")

    # ====================== ФУНКЦИИ ОТЧЁТОВ ======================
    def report_stock(self):
        conn = self.db.get_connection()

        # === 1. Детальный отчёт (по каждой партии) ===
        df_detail = pd.read_sql_query('''
                                      SELECT mp.name           as "Медчасть",
                                             m.name            as "Препарат",
                                             m.form            as "Форма",
                                             m.dosage          as "Дозировка",
                                             m.unit            as "Ед.",
                                             s.quantity        as "Количество",
                                             s.expiration_date as "Срок годности",
                                             s.batch_number    as "Серия",
                                             s.production_date as "Дата производства"
                                      FROM stock s
                                               JOIN med_parts mp ON s.med_part_id = mp.id
                                               JOIN medications m ON s.medication_id = m.id
                                      ORDER BY mp.name, m.name, s.expiration_date
                                      ''', conn)

        # === 2. Сводка — суммируем строго по Названию препарата + Медчасть ===
        df_summary = pd.read_sql_query('''
                                       SELECT mp.name                         as "Медчасть",
                                              m.name                          as "Препарат",
                                              GROUP_CONCAT(DISTINCT m.form)   as "Формы выпуска",
                                              GROUP_CONCAT(DISTINCT m.dosage) as "Дозировки",
                                              m.unit                          as "Ед.",
                                              SUM(s.quantity)                 as "Общее количество",
                                              COUNT(*)                        as "Кол-во партий",
                                              MIN(s.expiration_date)          as "Самый ранний срок",
                                              MAX(s.expiration_date)          as "Самый поздний срок"
                                       FROM stock s
                                                JOIN med_parts mp ON s.med_part_id = mp.id
                                                JOIN medications m ON s.medication_id = m.id
                                       GROUP BY mp.name, m.name, m.unit
                                       ORDER BY mp.name, m.name
                                       ''', conn)

        conn.close()

        # === Сохранение в Excel ===
        default_filename = f"Остатки_Медикаментов_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=default_filename,
            filetypes=[("Excel файлы", "*.xlsx")]
        )
        if not file_path:
            return

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df_detail.to_excel(writer, sheet_name="Детальные остатки", index=False)
            df_summary.to_excel(writer, sheet_name="Сводка по препаратам", index=False)

            # Красивое форматирование для обоих листов
            for sheet_name, df in [("Детальные остатки", df_detail), ("Сводка по препаратам", df_summary)]:
                ws = writer.sheets[sheet_name]

                # Автоширина
                for idx, col in enumerate(df.columns):
                    series = df[col].astype(str)
                    max_length = max(series.str.len().max() if not series.empty else 0, len(col)) + 3
                    ws.column_dimensions[chr(65 + idx)].width = min(max_length, 60)

                # Стили заголовков
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                alignment = Alignment(horizontal="center", vertical="center")

                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = alignment

                # Границы
                thin = Side(style='thin')
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                for row in ws.iter_rows():
                    for cell in row:
                        cell.border = border

        messagebox.showinfo("✅ Успешно", f"Отчёт сохранён!\n\n{file_path}\n(2 листа)")

    def report_writeoffs(self, date_from, date_to, part_filter):
        conn = self.db.get_connection()
        query = '''
            SELECT 
                DATE(w.date) as "Дата",
                mp.name as "Медчасть",
                m.name as "Препарат",
                w.quantity as "Кол-во",
                w.reason as "Причина",
                w.comment as "Комментарий"
            FROM writeoffs w
            JOIN stock s ON w.stock_id = s.id
            JOIN med_parts mp ON s.med_part_id = mp.id
            JOIN medications m ON s.medication_id = m.id
            WHERE DATE(w.date) BETWEEN ? AND ?
        '''
        params = [date_from, date_to]
        if part_filter != "Все медчасти":
            query += " AND mp.name = ?"
            params.append(part_filter)

        df = pd.read_sql_query(query, conn, params=params)
        conn.close()
        self._save_to_excel(df, "Списания_за_период")

    def report_movements(self, date_from, date_to, part_filter):
        conn = self.db.get_connection()
        query = '''
            SELECT 
                DATE(m.date) as "Дата",
                f.name as "Откуда",
                t.name as "Куда",
                med.name as "Препарат",
                m.quantity as "Кол-во",
                m.comment as "Комментарий"
            FROM movements m
            JOIN med_parts f ON m.from_part_id = f.id
            JOIN med_parts t ON m.to_part_id = t.id
            JOIN medications med ON m.medication_id = med.id
            WHERE DATE(m.date) BETWEEN ? AND ?
        '''
        params = [date_from, date_to]
        if part_filter != "Все медчасти":
            query += " AND (f.name = ? OR t.name = ?)"
            params.extend([part_filter, part_filter])

        df = pd.read_sql_query(query, conn, params=params)
        conn.close()
        self._save_to_excel(df, "Перемещения_за_период")

    def report_medical_records(self, date_from, date_to, prisoner_filter=None):
        """Отчёт по учёту медицинской помощи"""
        conn = self.db.get_connection()

        query = '''
                SELECT p.full_name   as "ФИО", \
                       p.prisoner_id as "№ дела", \
                       p.birth_date  as "Дата рожд.", DATE (mr.service_date) as "Дата выполн.", mr.diagnosis as "Диагноз", mr.doctor_name as "Врач", mr.service_name as "Наименование услуги", mr.quantity as "Кол-во", mr.unit_price as "Стоимость за ед.", mr.total_amount as "Сумма", mr.service_number as "Номер услуги", mr.notes as "Примечания"
                FROM medical_records mr
                    JOIN prisoners p \
                ON mr.prisoner_id = p.id
                WHERE DATE (mr.service_date) BETWEEN ? AND ? \
                '''
        params = [date_from, date_to]

        if prisoner_filter and prisoner_filter != "Все заключённые":
            # Извлекаем prisoner_id из строки вида "Иванов Иван Иванович (П-12345)"
            try:
                code = prisoner_filter.split('(')[-1].strip(')')
                query += " AND p.prisoner_id = ?"
                params.append(code)
            except:
                pass

        query += " ORDER BY mr.service_date DESC, p.full_name"

        df = pd.read_sql_query(query, conn, params=params)
        conn.close()

        self._save_to_excel(df, "Учет_Медицинской_Помощи")

    def _save_to_excel(self, df, base_name):
        if df.empty:
            messagebox.showinfo("Информация", "Нет данных для формирования отчёта")
            return

        default_filename = f"{base_name}_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=default_filename,
            filetypes=[("Excel файлы", "*.xlsx")]
        )
        if not file_path:
            return

        # === ПРЕОБРАЗОВАНИЕ ДАТ В ФОРМАТ ДД.ММ.ГГГГ ===
        date_keywords = ["дата", "срок", "произв", "прибыт", "рожд", "окончания"]
        for col in df.columns:
            col_lower = col.lower()
            if any(keyword in col_lower for keyword in date_keywords):
                df[col] = df[col].apply(lambda x: self.format_date(x) if pd.notna(x) and x != "" else x)

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name="Отчёт", index=False)
            ws = writer.sheets["Отчёт"]

            # === Автоширина колонок (исправлено) ===
            for idx, col in enumerate(df.columns):
                series = df[col].astype(str)
                # Защита от NaN и float
                max_length = max(
                    series.str.len().max() if not series.empty else 0,
                    len(col)
                ) + 3
                ws.column_dimensions[chr(65 + idx)].width = min(max_length, 60)

            # Стили заголовков
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            alignment = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment

            # Границы ячеек
            thin = Side(style='thin')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border

        messagebox.showinfo("✅ Успешно", f"Отчёт успешно сохранён!\n\n{file_path}")

    # ====================== СКЛАД ======================
    def create_stock_tab(self):
        top = ttk.Frame(self.tab_stock)
        top.pack(fill="x", padx=10, pady=8)

        ttk.Button(top, text="➕ Добавить медикамент", command=self.add_medication).pack(side=tk.LEFT, padx=5)
        ttk.Button(top, text="📉 Расходовать", command=self.write_off).pack(side=tk.LEFT, padx=5)  # ← Изменено
        ttk.Button(top, text="🔄 Переместить", command=self.move_medication).pack(side=tk.LEFT, padx=5)
        ttk.Button(top, text="🗑 Удалить", command=self.delete_stock_item).pack(side=tk.LEFT, padx=5)
        ttk.Button(top, text="🔄 Обновить", command=self.load_stock).pack(side=tk.LEFT, padx=5)

        ttk.Label(top, text="Поиск:").pack(side=tk.LEFT, padx=(30, 5))
        self.search_var = tk.StringVar()
        ttk.Entry(top, textvariable=self.search_var, width=30).pack(side=tk.LEFT, padx=5)
        self.search_var.trace_add("write", lambda *args: self.load_stock())

        ttk.Label(top, text="Медчасть:").pack(side=tk.LEFT, padx=(15, 5))
        self.part_var = tk.StringVar()
        self.part_combo = ttk.Combobox(top, textvariable=self.part_var, width=25, state="readonly")
        self.part_combo.pack(side=tk.LEFT, padx=5)
        self.part_combo.bind("<<ComboboxSelected>>", lambda e: self.load_stock())

        # Таблица с двумя датами
        cols = ("id", "part", "med", "qty", "unit", "prod", "exp", "batch")
        self.tree_stock = ttk.Treeview(self.tab_stock, columns=cols, show="headings", height=20)

        config = [
            ("ID", 55, "center"),
            ("Медчасть", 140, "w"),
            ("Препарат", 260, "w"),
            ("Кол-во", 95, "center"),
            ("Ед.", 65, "center"),
            ("Дата произв.", 110, "center"),
            ("Срок годн.", 110, "center"),
            ("Серия", 130, "w")
        ]

        for col, (text, width, anchor) in zip(cols, config):
            self.tree_stock.heading(col, text=text)
            self.tree_stock.column(col, width=width, anchor=anchor)

        scroll = ttk.Scrollbar(self.tab_stock, orient="vertical", command=self.tree_stock.yview)
        self.tree_stock.configure(yscrollcommand=scroll.set)
        self.tree_stock.pack(side=tk.LEFT, fill="both", expand=True, padx=(10, 0), pady=5)
        scroll.pack(side=tk.RIGHT, fill="y", pady=5)

        self.load_med_parts()
        self.load_stock()

    # ====================== ПЕРЕМЕЩЕНИЯ ======================
    def create_movements_tab(self):
        top = ttk.Frame(self.tab_movements)
        top.pack(fill="x", padx=10, pady=8)

        # Кнопки
        ttk.Button(top, text="🔄 Обновить", command=self.load_movements).pack(side=tk.LEFT, padx=5)
        ttk.Button(top, text="🗑 Удалить", command=self.delete_movement).pack(side=tk.LEFT, padx=5)

        # Поиск по препарату
        ttk.Label(top, text="Поиск препарата:").pack(side=tk.LEFT, padx=(25, 5))
        self.movement_search_var = tk.StringVar()
        ttk.Entry(top, textvariable=self.movement_search_var, width=28).pack(side=tk.LEFT, padx=5)
        self.movement_search_var.trace_add("write", lambda *args: self.load_movements())

        # Фильтр по медчасти
        ttk.Label(top, text="Медчасть:").pack(side=tk.LEFT, padx=(15, 5))
        self.movement_part_var = tk.StringVar()
        self.movement_part_combo = ttk.Combobox(top, textvariable=self.movement_part_var,
                                                width=20, state="readonly")
        self.movement_part_combo.pack(side=tk.LEFT, padx=5)
        self.movement_part_combo.bind("<<ComboboxSelected>>", lambda e: self.load_movements())

        # Таблица
        cols = ("id", "date", "from_part", "to_part", "med", "qty", "comment")
        self.tree_movements = ttk.Treeview(self.tab_movements, columns=cols, show="headings", height=22)
        headers = ["ID", "Дата", "Откуда", "Куда", "Препарат", "Кол-во", "Комментарий"]

        for col, text in zip(cols, headers):
            self.tree_movements.heading(col, text=text)
            if col == "id":
                self.tree_movements.column(col, width=60, anchor="center")
            elif col in ("qty", "date"):
                self.tree_movements.column(col, width=110, anchor="center")
            elif col == "comment":
                self.tree_movements.column(col, width=280)
            else:
                self.tree_movements.column(col, width=140)

        scroll = ttk.Scrollbar(self.tab_movements, orient="vertical", command=self.tree_movements.yview)
        self.tree_movements.configure(yscrollcommand=scroll.set)
        self.tree_movements.pack(side=tk.LEFT, fill="both", expand=True, padx=10, pady=5)
        scroll.pack(side=tk.RIGHT, fill="y", pady=5)

        self.load_med_parts_for_movements()
        self.load_movements()

    # ====================== РАСХОДОВАНИЕ ======================
    def create_expenditures_tab(self):
        top = ttk.Frame(self.tab_expenditures)
        top.pack(fill="x", padx=10, pady=8)

        ttk.Button(top, text="🔄 Обновить", command=self.load_expenditures).pack(side=tk.LEFT, padx=5)
        ttk.Button(top, text="🗑 Удалить", command=self.delete_expenditure).pack(side=tk.LEFT, padx=5)

        ttk.Label(top, text="Поиск препарата:").pack(side=tk.LEFT, padx=(25, 5))
        self.expenditure_search_var = tk.StringVar()
        ttk.Entry(top, textvariable=self.expenditure_search_var, width=30).pack(side=tk.LEFT, padx=5)
        self.expenditure_search_var.trace_add("write", lambda *args: self.load_expenditures())

        ttk.Label(top, text="Медчасть:").pack(side=tk.LEFT, padx=(15, 5))
        self.expenditure_part_var = tk.StringVar()
        self.expenditure_part_combo = ttk.Combobox(top, textvariable=self.expenditure_part_var,
                                                   width=22, state="readonly")
        self.expenditure_part_combo.pack(side=tk.LEFT, padx=5)
        self.expenditure_part_combo.bind("<<ComboboxSelected>>", lambda e: self.load_expenditures())

        # Таблица
        cols = ("id", "date", "part", "med", "prisoner", "qty", "reason", "comment")
        self.tree_expenditures = ttk.Treeview(self.tab_expenditures, columns=cols, show="headings", height=22)
        headers = ["ID", "Дата", "Медчасть", "Препарат", "Заключённый", "Кол-во", "Причина", "Комментарий"]

        for col, text in zip(cols, headers):
            self.tree_expenditures.heading(col, text=text)
            if col in ("id", "qty"):
                self.tree_expenditures.column(col, width=70, anchor="center")
            elif col == "date":
                self.tree_expenditures.column(col, width=110, anchor="center")
            elif col == "prisoner":
                self.tree_expenditures.column(col, width=220)
            elif col == "comment":
                self.tree_expenditures.column(col, width=280)
            else:
                self.tree_expenditures.column(col, width=130)

        scroll = ttk.Scrollbar(self.tab_expenditures, orient="vertical", command=self.tree_expenditures.yview)
        self.tree_expenditures.configure(yscrollcommand=scroll.set)
        self.tree_expenditures.pack(side=tk.LEFT, fill="both", expand=True, padx=10, pady=5)
        scroll.pack(side=tk.RIGHT, fill="y", pady=5)

        self.load_med_parts_for_expenditures()
        self.load_expenditures()
    # ====================== АМБУЛАТОРНАЯ ПОМОЩЬ ======================
    def create_prisoners_tab(self):
        top_frame = ttk.LabelFrame(self.tab_prisoners, text="Заключённые", padding=8)
        top_frame.pack(fill="x", padx=10, pady=5)

        btn_frame = ttk.Frame(top_frame)
        btn_frame.pack(fill="x", pady=(0, 8))
        ttk.Button(btn_frame, text="➕ Добавить заключённого", command=self.add_prisoner).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="📋 Новое обращение", command=self.add_medical_record).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="🗑 Удалить заключённого", command=self.delete_prisoner).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="🔄 Обновить", command=self.load_prisoners).pack(side=tk.LEFT, padx=5)

        ttk.Label(btn_frame, text="Поиск:").pack(side=tk.LEFT, padx=(20, 5))
        self.prisoner_search = tk.StringVar()
        ttk.Entry(btn_frame, textvariable=self.prisoner_search, width=35).pack(side=tk.LEFT, padx=5)
        self.prisoner_search.trace_add("write", lambda *args: self.load_prisoners())

        # Таблица заключённых
        cols1 = ("id", "p_id", "name", "birth", "arrival")
        self.tree_prisoners = ttk.Treeview(top_frame, columns=cols1, show="headings", height=10)
        h1 = ["ID", "№ дела", "ФИО", "Дата рожд.", "Прибытие"]
        for c, t in zip(cols1, h1):
            self.tree_prisoners.heading(c, text=t)
            if c == "id":
                self.tree_prisoners.column(c, width=60, anchor="center")
            elif c == "p_id":
                self.tree_prisoners.column(c, width=90, anchor="center")
            elif c == "name":
                self.tree_prisoners.column(c, width=320)
            else:
                self.tree_prisoners.column(c, width=110)

        sc1 = ttk.Scrollbar(top_frame, orient="vertical", command=self.tree_prisoners.yview)
        self.tree_prisoners.configure(yscrollcommand=sc1.set)
        self.tree_prisoners.pack(side=tk.LEFT, fill="both", expand=True, padx=(5, 0))
        sc1.pack(side=tk.RIGHT, fill="y")

        # История обращений
        bottom = ttk.LabelFrame(self.tab_prisoners, text="История обращений", padding=8)
        bottom.pack(fill="both", expand=True, padx=10, pady=8)

        # Кнопка удаления
        history_btn_frame = ttk.Frame(bottom)
        history_btn_frame.pack(fill="x", pady=(0, 6))
        ttk.Button(history_btn_frame, text="🗑 Удалить обращение",
                  command=self.delete_medical_record).pack(side=tk.LEFT, padx=5)

        cols2 = ("date", "service", "diagnosis", "doctor", "qty", "price", "number", "notes")
        self.tree_records = ttk.Treeview(bottom, columns=cols2, show="headings", height=12)
        h2 = ["Дата", "Услуга", "Диагноз", "Врач", "Кол-во", "Стоимость", "№ услуги", "Примечания"]
        for c, t in zip(cols2, h2):
            self.tree_records.heading(c, text=t)

        self.tree_records.column("date", width=100)
        self.tree_records.column("service", width=170)
        self.tree_records.column("diagnosis", width=140)
        self.tree_records.column("doctor", width=140)
        self.tree_records.column("qty", width=70, anchor="center")
        self.tree_records.column("price", width=90, anchor="center")
        self.tree_records.column("number", width=90, anchor="center")
        self.tree_records.column("notes", width=200)

        sc2 = ttk.Scrollbar(bottom, orient="vertical", command=self.tree_records.yview)
        self.tree_records.configure(yscrollcommand=sc2.set)
        self.tree_records.pack(side=tk.LEFT, fill="both", expand=True, padx=(5, 0))
        sc2.pack(side=tk.RIGHT, fill="y")

        self.tree_prisoners.bind("<<TreeviewSelect>>", self.show_prisoner_history)
        self.load_prisoners()

    def show_prisoner_history(self, event=None):
        """Показать историю обращений выбранного заключённого"""
        for item in self.tree_records.get_children():
            self.tree_records.delete(item)

        sel = self.tree_prisoners.selection()
        if not sel:
            return
        prisoner_id = self.tree_prisoners.item(sel[0])['values'][0]

        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
                       SELECT service_date, 
                              service_name, 
                              diagnosis, 
                              doctor_name,
                              quantity,
                              unit_price,
                              total_amount,
                              service_number,
                              notes
                       FROM medical_records
                       WHERE prisoner_id = ?
                       ORDER BY service_date DESC
                       ''', (prisoner_id,))

        for row in cursor.fetchall():
            date_only = self.format_date(str(row["service_date"])[:10] if row["service_date"] else None)

            self.tree_records.insert("", "end", values=(
                date_only,
                row["service_name"],
                row["diagnosis"] or "-",
                row["doctor_name"] or "-",
                f"{row['quantity']:.1f}" if row["quantity"] is not None else "1.0",
                f"{row['unit_price']:.2f}" if row["unit_price"] is not None else "0.00",   # ← Стоимость за единицу
                row["service_number"] or "-",
                str(row["notes"])[:60] if row["notes"] else "-"
            ))
        conn.close()

    # ====================== ВСПОМОГАТЕЛЬНЫЕ МЕТОДЫ ======================
    def load_med_parts(self):
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM med_parts ORDER BY name")
        self.part_combo['values'] = ["Все медчасти"] + [r["name"] for r in cursor.fetchall()]
        self.part_combo.current(0)
        conn.close()

    def format_date(self, date_str):
        """Преобразует YYYY-MM-DD в ДД.ММ.ГГГГ"""
        if not date_str or date_str == "None":
            return "-"
        try:
            if len(date_str) >= 10:
                y, m, d = date_str[:10].split('-')
                return f"{d}.{m}.{y}"
        except:
            pass
        return date_str

    # === НОВЫЙ МЕТОД ===
    def convert_date_for_query(self, date_str):
        """Преобразует ДД.ММ.ГГГГ → YYYY-MM-DD для SQL-запросов"""
        if not date_str or not date_str.strip():
            return None
        try:
            # Убираем лишние пробелы
            date_str = date_str.strip()
            day, month, year = date_str.split('.')
            return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
        except:
            # Если уже в формате YYYY-MM-DD или ошибка — возвращаем как есть
            return date_str

    def load_med_parts_for_movements(self):
        """Загрузка медчастей для фильтра в разделе Перемещения"""
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM med_parts ORDER BY name")
        parts = ["Все медчасти"] + [r["name"] for r in cursor.fetchall()]
        self.movement_part_combo['values'] = parts
        self.movement_part_combo.current(0)
        conn.close()

    def load_stock(self, *args):
        for item in self.tree_stock.get_children():
            self.tree_stock.delete(item)

        search = self.search_var.get().strip()
        part = self.part_var.get()

        conn = self.db.get_connection()
        cursor = conn.cursor()

        query = '''
                SELECT s.id,
                       mp.name as part,
                       m.name  as med,
                       s.quantity,
                       m.unit,
                       s.production_date,
                       s.expiration_date,
                       s.batch_number
                FROM stock s
                         JOIN med_parts mp ON s.med_part_id = mp.id
                         JOIN medications m ON s.medication_id = m.id
                WHERE 1 = 1
                '''
        params = []

        if part != "Все медчасти":
            query += " AND mp.name = ?"
            params.append(part)

        if search:
            search_lower = search.lower()
            query += " AND (LOWER(m.name) LIKE ? OR m.name LIKE ?)"
            params.extend([f"%{search_lower}%", f"%{search}%"])

        query += " ORDER BY mp.name, m.name"

        cursor.execute(query, params)

        for r in cursor.fetchall():
            prod = self.format_date(r["production_date"])
            exp = self.format_date(r["expiration_date"])

            self.tree_stock.insert("", "end", values=(
                r["id"],
                r["part"],
                r["med"],
                f"{r['quantity']:.2f}",
                r["unit"],
                prod,                       # Дата производства
                exp,                        # Срок годности (окончания)
                r["batch_number"] or "-"
            ))
        conn.close()
    def validate_date(self, new_text):
        """Строгая валидация + автоформатирование ДД.ММ.ГГГГ"""
        if not new_text:
            return True

        digits = ''.join(c for c in new_text if c.isdigit())
        if len(digits) > 8:
            return False

        formatted = ""
        if len(digits) >= 2:
            formatted = digits[:2]
        if len(digits) > 2:
            formatted += '.' + digits[2:4]
        if len(digits) > 4:
            formatted += '.' + digits[4:8]

        current = self.root.focus_get()
        if current and current.get() != formatted:
            current.delete(0, tk.END)
            current.insert(0, formatted)

        return True

    def load_movements(self):
        for item in self.tree_movements.get_children():
            self.tree_movements.delete(item)

        search = self.movement_search_var.get().strip()
        part_filter = self.movement_part_var.get() if hasattr(self, 'movement_part_var') else "Все медчасти"

        conn = self.db.get_connection()
        cursor = conn.cursor()

        query = '''
                SELECT m.id,
                       m.date,
                       f.name   as from_part,
                       t.name   as to_part,
                       med.name as medication,
                       m.quantity,
                       m.comment
                FROM movements m
                         JOIN med_parts f ON m.from_part_id = f.id
                         JOIN med_parts t ON m.to_part_id = t.id
                         JOIN medications med ON m.medication_id = med.id
                WHERE 1 = 1
                '''
        params = []

        if part_filter != "Все медчасти":
            query += " AND (f.name = ? OR t.name = ?)"
            params.extend([part_filter, part_filter])

        if search:
            search_lower = search.lower()
            query += " AND (LOWER(med.name) LIKE ? OR med.name LIKE ?)"
            params.extend([f"%{search_lower}%", f"%{search}%"])

        query += " ORDER BY m.date DESC"

        cursor.execute(query, params)
        for r in cursor.fetchall():
            date_only = self.format_date(str(r["date"])[:10])

            self.tree_movements.insert("", "end", values=(
                r["id"],
                date_only,
                r["from_part"],
                r["to_part"],
                r["medication"],
                f"{r['quantity']:.2f}",
                r["comment"] or "-"
            ))
        conn.close()

    def load_med_parts_for_writeoffs(self):
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM med_parts ORDER BY name")
        parts = ["Все медчасти"] + [r["name"] for r in cursor.fetchall()]
        self.writeoff_part_combo['values'] = parts
        self.writeoff_part_combo.current(0)
        conn.close()

    def load_writeoffs(self, *args):
        for item in self.tree_writeoffs.get_children():
            self.tree_writeoffs.delete(item)

        search = self.writeoff_search_var.get().strip().lower() if hasattr(self, 'writeoff_search_var') else ""
        part_filter = self.writeoff_part_var.get() if hasattr(self, 'writeoff_part_var') else "Все медчасти"

        conn = self.db.get_connection()
        cursor = conn.cursor()

        query = '''
            SELECT w.id,
                   w.date,
                   mp.name as part,
                   m.name as medication,
                   w.quantity,
                   w.reason,
                   w.comment
            FROM writeoffs w
            JOIN stock s ON w.stock_id = s.id
            JOIN med_parts mp ON s.med_part_id = mp.id
            JOIN medications m ON s.medication_id = m.id
            WHERE 1 = 1
        '''
        params = []

        if part_filter != "Все медчасти":
            query += " AND mp.name = ?"
            params.append(part_filter)

        if search:
            query += " AND LOWER(m.name) LIKE ?"
            params.append(f"%{search}%")

        query += " ORDER BY w.date DESC"

        cursor.execute(query, params)
        for r in cursor.fetchall():
            date_only = self.format_date(str(r["date"])[:10])   # ← Только дата

            self.tree_writeoffs.insert("", "end", values=(
                r["id"],
                date_only,                    # ← Красивый формат ДД.ММ.ГГГГ
                r["part"],
                r["medication"],
                f"{r['quantity']:.2f}",
                r["reason"],
                r["comment"] or "-"
            ))
        conn.close()

    def delete_writeoff(self):
        sel = self.tree_writeoffs.selection()
        if not sel:
            messagebox.showwarning("Внимание", "Выберите списание для удаления")
            return

        values = self.tree_writeoffs.item(sel[0])['values']
        w_id = values[0]
        date = values[1]
        med = values[3]
        qty = values[4]

        if not messagebox.askyesno("Подтверждение удаления",
                                   f"Удалить списание?\n\n"
                                   f"Дата: {date}\n"
                                   f"Препарат: {med}\n"
                                   f"Количество: {qty}"):
            return

        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM writeoffs WHERE id = ?", (w_id,))
            conn.commit()
            conn.close()

            messagebox.showinfo("Успешно", "Списание удалено")
            self.load_writeoffs()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось удалить:\n{str(e)}")


    def load_prisoners(self, *args):
        for item in self.tree_prisoners.get_children():
            self.tree_prisoners.delete(item)

        search = self.prisoner_search.get().strip().lower()
        conn = self.db.get_connection()
        cursor = conn.cursor()

        if search:
            cursor.execute('''
                           SELECT id, prisoner_id, full_name, birth_date, arrival_date
                           FROM prisoners
                           WHERE LOWER(full_name) LIKE ?
                              OR prisoner_id LIKE ?
                           ''', (f"%{search}%", f"%{search}%"))
        else:
            cursor.execute("SELECT id, prisoner_id, full_name, birth_date, arrival_date FROM prisoners")

        for row in cursor.fetchall():
            self.tree_prisoners.insert("", "end", values=(
                row["id"],
                row["prisoner_id"],
                row["full_name"],
                self.format_date(row["birth_date"]),
                self.format_date(row["arrival_date"])
            ))
        conn.close()

    def delete_movement(self):
        """Удаление перемещения"""
        sel = self.tree_movements.selection()
        if not sel:
            messagebox.showwarning("Внимание", "Выберите перемещение для удаления")
            return

        values = self.tree_movements.item(sel[0])['values']
        mov_id = values[0]
        date = values[1]
        med = values[4]

        if not messagebox.askyesno("Подтверждение удаления",
                                   f"Удалить перемещение?\n\nДата: {date}\nПрепарат: {med}"):
            return

        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM movements WHERE id = ?", (mov_id,))
            conn.commit()
            conn.close()

            messagebox.showinfo("Успешно", "Перемещение удалено")
            self.load_movements()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось удалить:\n{str(e)}")

    def delete_medical_record(self):
        """Удаление обращения — улучшенная версия"""
        sel = self.tree_records.selection()
        if not sel:
            messagebox.showwarning("Внимание", "Выберите обращение для удаления")
            return

        values = self.tree_records.item(sel[0])['values']
        displayed_date = values[0]   # Например: "05.06.2025"
        service_name = values[1]

        prisoner_sel = self.tree_prisoners.selection()
        if not prisoner_sel:
            messagebox.showwarning("Внимание", "Сначала выберите заключённого")
            return

        prisoner_id = self.tree_prisoners.item(prisoner_sel[0])['values'][0]

        if not messagebox.askyesno("Подтверждение удаления",
                                   f"Удалить обращение?\n\n"
                                   f"Дата: {displayed_date}\n"
                                   f"Услуга: {service_name}"):
            return

        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()

            # Более надёжный способ удаления — по prisoner_id + service_name + дата (с диапазоном)
            # Преобразуем отображаемую дату в SQL-формат
            try:
                day, month, year = displayed_date.split('.')
                date_start = f"{year}-{month}-{day} 00:00:00"
                date_end = f"{year}-{month}-{day} 23:59:59"
            except:
                date_start = displayed_date
                date_end = displayed_date

            cursor.execute('''
                DELETE FROM medical_records
                WHERE prisoner_id = ?
                  AND service_name = ?
                  AND service_date >= ? 
                  AND service_date <= ?
            ''', (prisoner_id, service_name, date_start, date_end))

            deleted_rows = cursor.rowcount
            conn.commit()
            conn.close()

            if deleted_rows > 0:
                messagebox.showinfo("Успешно", f"Обращение успешно удалено ({deleted_rows} записей)")
            else:
                # Дополнительная попытка — только по prisoner_id и service_name
                conn = self.db.get_connection()
                cursor = conn.cursor()
                cursor.execute('''
                    DELETE FROM medical_records
                    WHERE prisoner_id = ? AND service_name = ?
                ''', (prisoner_id, service_name))
                deleted_rows2 = cursor.rowcount
                conn.commit()
                conn.close()

                if deleted_rows2 > 0:
                    messagebox.showinfo("Успешно", "Обращение удалено (по упрощённому поиску)")
                else:
                    messagebox.showwarning("Предупреждение", "Запись не найдена или уже удалена")

            self.show_prisoner_history()

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось удалить:\n{str(e)}")

    def add_medication(self):
        AddMedicationWindow(self.db, self.load_stock)

    def write_off(self):
        sel = self.tree_stock.selection()
        if not sel: return messagebox.showwarning("Внимание", "Выберите медикамент")
        v = self.tree_stock.item(sel[0])['values']
        WriteOffWindow(self.db, v[0], v[2], float(v[3]), self.load_stock)

    def move_medication(self):
        sel = self.tree_stock.selection()
        if not sel: return messagebox.showwarning("Внимание", "Выберите медикамент")
        v = self.tree_stock.item(sel[0])['values']
        MoveWindow(self.db, v[0], v[1], v[2], float(v[3]), self.load_stock)

    def add_prisoner(self):
        AddPrisonerWindow(self.db, self.load_prisoners)

    def add_medical_record(self):
        sel = self.tree_prisoners.selection()
        if not sel:
            messagebox.showwarning("Внимание", "Сначала выберите заключённого")
            return
        pid = self.tree_prisoners.item(sel[0])['values'][0]
        name = self.tree_prisoners.item(sel[0])['values'][2]
        AddMedicalRecordWindow(self.db, pid, name, self.show_prisoner_history)

    def run(self):
        self.root.mainloop()

    def delete_prisoner(self):
        """Удаление заключённого"""
        sel = self.tree_prisoners.selection()
        if not sel:
            messagebox.showwarning("Внимание", "Выберите заключённого для удаления")
            return

        values = self.tree_prisoners.item(sel[0])['values']
        prisoner_id = values[0]
        full_name = values[2]

        if not messagebox.askyesno("ВНИМАНИЕ! УДАЛЕНИЕ ЗАКЛЮЧЁННОГО",
                                   f"Вы действительно хотите удалить:\n\n"
                                   f"{full_name}\n"
                                   f"№ дела: {values[1]}\n\n"
                                   f"Вместе с ним удалятся все его медицинские обращения!\n"
                                   f"Это действие нельзя отменить!"):
            return

        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()

            # Сначала удаляем все обращения
            cursor.execute("DELETE FROM medical_records WHERE prisoner_id = ?", (prisoner_id,))
            # Затем самого заключённого
            cursor.execute("DELETE FROM prisoners WHERE id = ?", (prisoner_id,))

            conn.commit()
            conn.close()

            messagebox.showinfo("Успешно", f"Заключённый {full_name} удалён")
            self.load_prisoners()
            # Очищаем историю
            for item in self.tree_records.get_children():
                self.tree_records.delete(item)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось удалить:\n{str(e)}")
    def load_med_parts_for_expenditures(self):
        """Загрузка медчастей для фильтра в разделе Расходование"""
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM med_parts ORDER BY name")
        parts = ["Все медчасти"] + [r["name"] for r in cursor.fetchall()]
        self.expenditure_part_combo['values'] = parts
        self.expenditure_part_combo.current(0)
        conn.close()

    def load_expenditures(self, *args):
        for item in self.tree_expenditures.get_children():
            self.tree_expenditures.delete(item)

        search = self.expenditure_search_var.get().strip()
        part_filter = self.expenditure_part_var.get() if hasattr(self, 'expenditure_part_var') else "Все медчасти"

        conn = self.db.get_connection()
        cursor = conn.cursor()

        query = '''
                SELECT e.id,
                       e.date,
                       mp.name                    as part,
                       m.name                     as medication,
                       COALESCE(p.full_name, '—') as prisoner,
                       e.quantity,
                       e.reason,
                       e.comment
                FROM expenditures e
                         JOIN stock s ON e.stock_id = s.id
                         JOIN med_parts mp ON s.med_part_id = mp.id
                         JOIN medications m ON s.medication_id = m.id
                         LEFT JOIN prisoners p ON e.prisoner_id = p.id
                WHERE 1 = 1
                '''
        params = []

        if part_filter != "Все медчасти":
            query += " AND mp.name = ?"
            params.append(part_filter)

        if search:
            search_lower = search.lower()
            query += " AND (LOWER(m.name) LIKE ? OR m.name LIKE ?)"
            params.extend([f"%{search_lower}%", f"%{search}%"])

        query += " ORDER BY e.date DESC"

        cursor.execute(query, params)
        for r in cursor.fetchall():
            date_only = self.format_date(str(r["date"])[:10])

            self.tree_expenditures.insert("", "end", values=(
                r["id"],
                date_only,
                r["part"],
                r["medication"],
                r["prisoner"],
                f"{r['quantity']:.2f}",
                r["reason"] or "—",
                r["comment"] or "-"
            ))
        conn.close()

    def delete_expenditure(self):
        """Удаление записи расходования"""
        sel = self.tree_expenditures.selection()
        if not sel:
            messagebox.showwarning("Внимание", "Выберите запись расходования для удаления")
            return

        values = self.tree_expenditures.item(sel[0])['values']
        exp_id = values[0]
        date = values[1]
        med = values[3]

        if not messagebox.askyesno("Подтверждение удаления",
                                  f"Удалить запись расходования?\n\n"
                                  f"Дата: {date}\n"
                                  f"Препарат: {med}"):
            return

        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM expenditures WHERE id = ?", (exp_id,))
            conn.commit()
            conn.close()

            messagebox.showinfo("Успешно", "Запись расходования удалена")
            self.load_expenditures()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось удалить:\n{str(e)}")

# ====================== ОКНА ======================
class AddMedicationWindow:
    def __init__(self, db, refresh):
        self.db = db
        self.refresh = refresh
        self.win = tk.Toplevel()
        self.win.title("Добавить медикамент на склад")
        self.win.geometry("550x580")
        self.win.grab_set()

        ttk.Label(self.win, text="Добавление медикамента", font=("Arial", 12, "bold")).pack(pady=10)
        ttk.Label(self.win, text="Медчасть: Медчасть №1 (центральный склад)",
                 font=("Arial", 10), foreground="blue").pack(pady=(0, 10))

        labels = ["Название препарата*", "Форма выпуска", "Дозировка", "Единица измерения*",
                  "Количество*", "Дата производства", "Дата окончания срока годности*",
                  "Серия"]

        self.entries = {}
        for label in labels:
            ttk.Label(self.win, text=label).pack(anchor="w", padx=30, pady=(8 if label.startswith("Название") else 5))

            entry = ttk.Entry(self.win, width=50)
            entry.pack(padx=30, fill="x")
            self.entries[label] = entry

            if "дата" in label.lower():
                vcmd = (self.win.register(self.validate_date), '%P')
                entry.configure(validate="key", validatecommand=vcmd)

        btn_frame = ttk.Frame(self.win)
        btn_frame.pack(pady=25)
        ttk.Button(btn_frame, text="Сохранить", command=self.save).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="Отмена", command=self.win.destroy).pack(side=tk.LEFT, padx=10)

    def validate_date(self, new_text):
        """Строгая валидация + автоформатирование ДД.ММ.ГГГГ"""
        if not new_text:
            return True

        # Оставляем только цифры
        digits = ''.join(c for c in new_text if c.isdigit())

        # Ограничение длины
        if len(digits) > 8:
            return False

        # Форматируем для отображения
        formatted = ""
        if len(digits) >= 2:
            formatted = digits[:2]
        if len(digits) > 2:
            formatted += '.' + digits[2:4]
        if len(digits) > 4:
            formatted += '.' + digits[4:8]

        # === СТРОГАЯ ПРОВЕРКА ===
        if len(digits) >= 4:  # когда введены день и месяц
            try:
                day = int(digits[:2])
                month = int(digits[2:4])

                if day < 1 or day > 31:
                    return False
                if month < 1 or month > 12:
                    return False

                # Дополнительно: для апреля, июня, сентября, ноября — max 30 дней
                if month in [4, 6, 9, 11] and day > 30:
                    return False
                # Для февраля — грубо max 29
                if month == 2 and day > 29:
                    return False

            except:
                return False

        # Обновляем поле
        current = self.win.focus_get()
        if current and current.get() != formatted:
            current.delete(0, tk.END)
            current.insert(0, formatted)

        return True

    def load_parts(self):
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM med_parts ORDER BY name")
        self.part_combo['values'] = [row["name"] for row in cursor.fetchall()]
        if self.part_combo['values']:
            self.part_combo.current(0)
        conn.close()

    def save(self):
        try:
            name = self.entries["Название препарата*"].get().strip()
            if not name:
                messagebox.showerror("Ошибка", "Название препарата обязательно!")
                return

            quantity = float(self.entries["Количество*"].get().strip() or 0)
            if quantity <= 0:
                messagebox.showerror("Ошибка", "Количество должно быть больше 0")
                return

            form = self.entries["Форма выпуска"].get().strip()
            dosage = self.entries["Дозировка"].get().strip()
            unit = self.entries["Единица измерения*"].get().strip() or "шт"
            batch = self.entries["Серия"].get().strip() or None

            # Преобразование дат
            prod_input = self.entries["Дата производства"].get().strip()
            exp_input = self.entries["Дата окончания срока годности*"].get().strip()

            production_date = self.convert_date(prod_input)
            expiration_date = self.convert_date(exp_input)

            conn = self.db.get_connection()
            cursor = conn.cursor()

            # Добавляем/обновляем препарат
            cursor.execute('''
                INSERT OR IGNORE INTO medications (name, form, dosage, unit)
                VALUES (?, ?, ?, ?)
            ''', (name, form, dosage, unit))

            cursor.execute("SELECT id FROM medications WHERE name = ? AND form = ? AND dosage = ? AND unit = ?",
                           (name, form, dosage, unit))
            med_id = cursor.fetchone()["id"]

            # Жёстко добавляем только в Медчасть №1
            cursor.execute("SELECT id FROM med_parts WHERE name = 'Медчасть №1'")
            part_result = cursor.fetchone()
            if not part_result:
                messagebox.showerror("Ошибка", "Медчасть №1 не найдена в базе!")
                conn.close()
                return
            part_id = part_result["id"]

            # Добавление на склад
            cursor.execute('''
                INSERT INTO stock
                (med_part_id, medication_id, quantity, production_date, expiration_date, batch_number)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (part_id, med_id, quantity, production_date, expiration_date, batch))

            conn.commit()
            conn.close()

            messagebox.showinfo("Успех", f"Препарат '{name}' добавлен на Медчасть №1")
            self.refresh()
            self.win.destroy()

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить:\n{str(e)}")

    def convert_date(self, date_str):
        if not date_str:
            return None
        try:
            day, month, year = date_str.split('.')
            return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
        except:
            return date_str


class WriteOffWindow:
    def __init__(self, db, stock_id, med_name, current_qty, refresh):
        self.db = db
        self.stock_id = stock_id
        self.med_name = med_name
        self.current_qty = current_qty
        self.refresh = refresh

        self.win = tk.Toplevel()
        self.win.title("Расход медикамента")
        self.win.geometry("560x480")
        self.win.grab_set()
        self.win.resizable(True, True)

        ttk.Label(self.win, text="Расход медикамента", font=("Arial", 12, "bold")).pack(pady=10)
        ttk.Label(self.win, text=f"Препарат: {med_name}", font=("Arial", 11)).pack(pady=5)
        ttk.Label(self.win, text=f"Текущее количество: {current_qty}").pack(pady=5)

        # Выбор заключённого
        ttk.Label(self.win, text="Заключённый:").pack(anchor="w", padx=30, pady=(15, 5))
        self.prisoner_combo = ttk.Combobox(self.win, width=50, state="readonly")
        self.prisoner_combo.pack(padx=30, fill="x")
        self.load_prisoners()

        ttk.Label(self.win, text="Количество для расхода:").pack(anchor="w", padx=30, pady=(15, 5))
        self.qty_var = tk.StringVar()
        ttk.Entry(self.win, textvariable=self.qty_var, width=20).pack(padx=30, fill="x")

        ttk.Label(self.win, text="Комментарий:").pack(anchor="w", padx=30, pady=(15, 5))
        self.comment_text = tk.Text(self.win, height=6)
        self.comment_text.pack(padx=30, fill="x", pady=5)

        btn_frame = ttk.Frame(self.win)
        btn_frame.pack(pady=25)
        ttk.Button(btn_frame, text="Расходовать", command=self.save).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="Отмена", command=self.win.destroy).pack(side=tk.LEFT, padx=10)

    def load_prisoners(self):
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, full_name, prisoner_id FROM prisoners ORDER BY full_name")
        prisoners = [f"{row['full_name']} ({row['prisoner_id']})" for row in cursor.fetchall()]
        self.prisoner_combo['values'] = prisoners
        if prisoners:
            self.prisoner_combo.current(0)
        conn.close()

    def save(self):
        try:
            qty = float(self.qty_var.get())
            if qty <= 0:
                messagebox.showerror("Ошибка", "Количество должно быть больше 0")
                return
            if qty > self.current_qty:
                messagebox.showerror("Ошибка", f"Нельзя расходовать больше чем есть ({self.current_qty})")
                return

            selected = self.prisoner_combo.get()
            if not selected:
                messagebox.showerror("Ошибка", "Выберите заключённого")
                return

            # Извлекаем настоящий ID заключённого (числовой id)
            prisoner_id = None
            if '(' in selected and ')' in selected:
                try:
                    # Берем часть в скобках (П-12345) и ищем по нему настоящий id
                    prisoner_code = selected.split('(')[-1].strip(')')
                    conn = self.db.get_connection()
                    cursor = conn.cursor()
                    cursor.execute("SELECT id FROM prisoners WHERE prisoner_id = ?", (prisoner_code,))
                    result = cursor.fetchone()
                    if result:
                        prisoner_id = result["id"]
                    conn.close()
                except:
                    prisoner_id = None

            comment = self.comment_text.get("1.0", tk.END).strip()

            conn = self.db.get_connection()
            cursor = conn.cursor()

            cursor.execute('''
                INSERT INTO expenditures (stock_id, prisoner_id, quantity, reason, comment)
                VALUES (?, ?, ?, ?, ?)
            ''', (self.stock_id, prisoner_id, qty, "Использовано при лечении", comment))

            cursor.execute("UPDATE stock SET quantity = quantity - ? WHERE id = ?",
                         (qty, self.stock_id))

            conn.commit()
            conn.close()

            messagebox.showinfo("Успех", f"Расходовано {qty} ед. препарата '{self.med_name}'")
            self.refresh()
            self.win.destroy()

        except ValueError:
            messagebox.showerror("Ошибка", "Введите корректное число")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить расход:\n{str(e)}")

class MoveWindow:
    def __init__(self, db, stock_id, current_part, med_name, current_qty, refresh):
        self.db = db
        self.stock_id = stock_id
        self.current_part = current_part
        self.med_name = med_name
        self.current_qty = current_qty
        self.refresh = refresh

        self.win = tk.Toplevel()
        self.win.title("Перемещение медикамента")
        self.win.geometry("500x400")
        self.win.grab_set()

        ttk.Label(self.win, text="Перемещение медикамента", font=("Arial", 12, "bold")).pack(pady=10)
        ttk.Label(self.win, text=f"Препарат: {med_name}").pack(pady=5)
        ttk.Label(self.win, text=f"Откуда: {current_part}").pack(pady=5)

        ttk.Label(self.win, text="Куда переместить:").pack(anchor="w", padx=30, pady=(15, 5))
        self.to_part_combo = ttk.Combobox(self.win, width=40, state="readonly")
        self.to_part_combo.pack(padx=30, fill="x")
        self.load_to_parts()

        ttk.Label(self.win, text="Количество для перемещения:").pack(anchor="w", padx=30, pady=(15, 5))
        self.qty_var = tk.StringVar()
        ttk.Entry(self.win, textvariable=self.qty_var, width=20).pack(padx=30, fill="x")

        ttk.Label(self.win, text="Комментарий:").pack(anchor="w", padx=30, pady=(10, 5))
        self.comment_text = tk.Text(self.win, height=4)
        self.comment_text.pack(padx=30, fill="x", pady=5)

        btn_frame = ttk.Frame(self.win)
        btn_frame.pack(pady=20)
        ttk.Button(btn_frame, text="Переместить", command=self.save).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="Отмена", command=self.win.destroy).pack(side=tk.LEFT, padx=10)

    def load_to_parts(self):
        conn = self.db.get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM med_parts WHERE name != ?", (self.current_part,))
        self.to_part_combo['values'] = [row["name"] for row in cursor.fetchall()]
        if self.to_part_combo['values']:
            self.to_part_combo.current(0)
        conn.close()

    def save(self):
        try:
            qty = float(self.qty_var.get())
            if qty <= 0 or qty > self.current_qty:
                messagebox.showerror("Ошибка", f"Количество должно быть от 0 до {self.current_qty}")
                return

            to_part = self.to_part_combo.get()
            if not to_part:
                messagebox.showerror("Ошибка", "Выберите медчасть назначения")
                return

            comment = self.comment_text.get("1.0", tk.END).strip()

            conn = self.db.get_connection()
            cursor = conn.cursor()

            # Получаем все данные исходной записи
            cursor.execute("""
                SELECT medication_id, med_part_id, production_date, expiration_date, batch_number 
                FROM stock WHERE id=?
            """, (self.stock_id,))
            stock_info = cursor.fetchone()

            medication_id = stock_info["medication_id"]
            from_part_id = stock_info["med_part_id"]
            production_date = stock_info["production_date"]
            expiration_date = stock_info["expiration_date"]
            batch_number = stock_info["batch_number"]

            # Получаем id целевой медчасти
            cursor.execute("SELECT id FROM med_parts WHERE name=?", (to_part,))
            to_part_id = cursor.fetchone()["id"]

            # Записываем перемещение
            cursor.execute('''
                INSERT INTO movements 
                (from_part_id, to_part_id, medication_id, quantity, comment)
                VALUES (?, ?, ?, ?, ?)
            ''', (from_part_id, to_part_id, medication_id, qty, comment))

            # Уменьшаем количество на исходной медчасти
            cursor.execute("UPDATE stock SET quantity = quantity - ? WHERE id = ?",
                         (qty, self.stock_id))

            # Добавляем/увеличиваем на целевой медчасти с сохранением дат
            cursor.execute('''
                SELECT id, quantity
                FROM stock
                WHERE med_part_id = ? 
                  AND medication_id = ?
                  AND (batch_number = ? OR (batch_number IS NULL AND ? IS NULL))
                  AND (expiration_date = ? OR (expiration_date IS NULL AND ? IS NULL))
            ''', (to_part_id, medication_id, batch_number, batch_number,
                  expiration_date, expiration_date))

            existing = cursor.fetchone()

            if existing:
                # Увеличиваем существующую партию
                cursor.execute("UPDATE stock SET quantity = quantity + ? WHERE id = ?",
                               (qty, existing["id"]))
            else:
                # Создаём новую запись с сохранением всех дат
                cursor.execute('''
                    INSERT INTO stock 
                    (med_part_id, medication_id, quantity, production_date, expiration_date, batch_number)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (to_part_id, medication_id, qty, production_date, expiration_date, batch_number))

            conn.commit()
            conn.close()

            messagebox.showinfo("Успех", f"Перемещено {qty} ед. в {to_part}")
            self.refresh()
            self.win.destroy()

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось выполнить перемещение:\n{str(e)}")


class AddPrisonerWindow:
    def __init__(self, db, refresh):
        self.db = db
        self.refresh = refresh
        self.win = tk.Toplevel()
        self.win.title("Добавить заключённого")
        self.win.geometry("500x480")
        self.win.grab_set()

        fields = ["Номер дела*", "ФИО*", "Дата рождения", "Дата прибытия", "Примечания"]
        self.entries = {}

        for f in fields:
            ttk.Label(self.win, text=f).pack(anchor="w", padx=30, pady=5)
            e = ttk.Entry(self.win, width=50)
            e.pack(padx=30, fill="x")
            self.entries[f] = e

            # Автоформатирование для дат
            if "Дата" in f:
                vcmd = (self.win.register(self.validate_date), '%P')
                e.configure(validate="key", validatecommand=vcmd)

        ttk.Button(self.win, text="Сохранить", command=self.save).pack(pady=20)

    def validate_date(self, new_text):
        """Строгая валидация + автоформатирование ДД.ММ.ГГГГ"""
        if not new_text:
            return True

        # Оставляем только цифры
        digits = ''.join(c for c in new_text if c.isdigit())

        # Ограничение длины
        if len(digits) > 8:
            return False

        # Форматируем для отображения
        formatted = ""
        if len(digits) >= 2:
            formatted = digits[:2]
        if len(digits) > 2:
            formatted += '.' + digits[2:4]
        if len(digits) > 4:
            formatted += '.' + digits[4:8]

        # === СТРОГАЯ ПРОВЕРКА ===
        if len(digits) >= 4:  # когда введены день и месяц
            try:
                day = int(digits[:2])
                month = int(digits[2:4])

                if day < 1 or day > 31:
                    return False
                if month < 1 or month > 12:
                    return False

                # Дополнительно: для апреля, июня, сентября, ноября — max 30 дней
                if month in [4, 6, 9, 11] and day > 30:
                    return False
                # Для февраля — грубо max 29
                if month == 2 and day > 29:
                    return False

            except:
                return False

        # Обновляем поле
        current = self.win.focus_get()
        if current and current.get() != formatted:
            current.delete(0, tk.END)
            current.insert(0, formatted)

        return True

    def save(self):
        try:
            conn = self.db.get_connection()
            cursor = conn.cursor()

            birth_input = self.entries["Дата рождения"].get().strip()
            arrival_input = self.entries["Дата прибытия"].get().strip()

            birth_date = self.convert_date(birth_input)
            arrival_date = self.convert_date(arrival_input)

            cursor.execute('''
                           INSERT INTO prisoners
                               (prisoner_id, full_name, birth_date, arrival_date, notes)
                           VALUES (?, ?, ?, ?, ?)
                           ''', (
                               self.entries["Номер дела*"].get().strip(),
                               self.entries["ФИО*"].get().strip(),
                               birth_date,
                               arrival_date,
                               self.entries["Примечания"].get().strip()
                           ))

            conn.commit()
            conn.close()

            messagebox.showinfo("Успех", "Заключённый добавлен")
            self.refresh()
            self.win.destroy()
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    def convert_date(self, date_str):
        """Преобразует ДД.ММ.ГГГГ в YYYY-MM-DD"""
        if not date_str:
            return None
        try:
            day, month, year = date_str.split('.')
            return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
        except:
            return date_str


class AddMedicalRecordWindow:
    def __init__(self, db, prisoner_id, prisoner_name, refresh):
        self.db = db
        self.prisoner_id = prisoner_id
        self.refresh = refresh
        self.win = tk.Toplevel()
        self.win.title("Новое обращение / Услуга")
        self.win.geometry("650x720")
        self.win.grab_set()

        ttk.Label(self.win, text=f"Заключённый: {prisoner_name}",
                  font=("Arial", 11, "bold")).pack(pady=10)

        # Дата выполнения
        ttk.Label(self.win, text="Дата выполнения:").pack(anchor="w", padx=30, pady=(10, 5))
        self.service_date = ttk.Entry(self.win, width=20)
        self.service_date.pack(padx=30, anchor="w")
        self.service_date.insert(0, datetime.now().strftime("%d.%m.%Y"))

        # Диагноз
        ttk.Label(self.win, text="Диагноз (МКБ-10):").pack(anchor="w", padx=30, pady=(15, 5))
        self.diagnosis_var = tk.StringVar()
        self.diagnosis_combo = ttk.Combobox(self.win, textvariable=self.diagnosis_var,
                                            width=70, state="normal")
        self.diagnosis_combo.pack(padx=30, fill="x")

        self.load_diagnoses()

        # Привязываем фильтрацию при каждом нажатии клавиши
        self.diagnosis_combo.bind("<KeyRelease>", self.filter_diagnoses)

        # Врач
        ttk.Label(self.win, text="Врач (ФИО):").pack(anchor="w", padx=30, pady=(15, 5))
        self.doctor_name = ttk.Entry(self.win, width=50)
        self.doctor_name.pack(padx=30, fill="x")
        self.doctor_name.insert(0, "Иванов И.И.")

        # Услуга
        ttk.Label(self.win, text="Наименование услуги:").pack(anchor="w", padx=30, pady=(10, 5))
        self.service_name = ttk.Entry(self.win, width=50)
        self.service_name.pack(padx=30, fill="x")

        # Количество и стоимость
        qty_frame = ttk.Frame(self.win)
        qty_frame.pack(fill="x", padx=30, pady=8)
        ttk.Label(qty_frame, text="Количество:").pack(side=tk.LEFT)
        self.quantity_var = tk.StringVar(value="1")
        ttk.Entry(qty_frame, textvariable=self.quantity_var, width=10).pack(side=tk.LEFT, padx=10)

        ttk.Label(qty_frame, text="Стоимость за единицу (руб):").pack(side=tk.LEFT, padx=(20, 5))
        self.unit_price_var = tk.StringVar(value="0.00")
        ttk.Entry(qty_frame, textvariable=self.unit_price_var, width=12).pack(side=tk.LEFT)

        # Номер услуги
        ttk.Label(self.win, text="Номер услуги:").pack(anchor="w", padx=30, pady=(10, 5))
        self.service_number = ttk.Entry(self.win, width=25)
        self.service_number.pack(padx=30, anchor="w")

        # Примечания
        ttk.Label(self.win, text="Примечания:").pack(anchor="w", padx=30, pady=(10, 5))
        self.notes = tk.Text(self.win, height=5)
        self.notes.pack(padx=30, fill="x")

        btn_frame = ttk.Frame(self.win)
        btn_frame.pack(pady=25)
        ttk.Button(btn_frame, text="Сохранить услугу", command=self.save).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="Отмена", command=self.win.destroy).pack(side=tk.LEFT, padx=10)

    def load_diagnoses(self):
        """Загружает диагнозы из файла"""
        try:
            with open("diagnoses.txt", "r", encoding="utf-8") as f:
                self.diagnoses = [line.strip() for line in f if line.strip() and not line.startswith("#")]

            if self.diagnoses:
                self.diagnosis_combo['values'] = self.diagnoses
            else:
                self.diagnosis_combo['values'] = ["Список диагнозов пуст"]
        except FileNotFoundError:
            messagebox.showwarning("Предупреждение",
                                   "Файл diagnoses.txt не найден.\nСоздайте его в корне проекта.")
            self.diagnoses = ["Файл diagnoses.txt не найден"]
            self.diagnosis_combo['values'] = self.diagnoses
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить диагнозы:\n{str(e)}")
            self.diagnoses = ["Ошибка загрузки"]
            self.diagnosis_combo['values'] = self.diagnoses

    def filter_diagnoses(self, event=None):
        """Фильтрация диагнозов при вводе текста"""
        typed = self.diagnosis_var.get().strip().lower()

        if not typed:
            self.diagnosis_combo['values'] = self.diagnoses
            return

        # Фильтруем по содержимому
        filtered = [d for d in self.diagnoses if typed in d.lower()]
        self.diagnosis_combo['values'] = filtered if filtered else self.diagnoses

    def validate_date(self, new_text):
        if not new_text:
            return True
        digits = ''.join(c for c in new_text if c.isdigit())
        if len(digits) > 8:
            return False

        formatted = ""
        if len(digits) >= 2: formatted = digits[:2]
        if len(digits) > 2: formatted += '.' + digits[2:4]
        if len(digits) > 4: formatted += '.' + digits[4:8]

        current = self.win.focus_get()
        if current and current.get() != formatted:
            current.delete(0, tk.END)
            current.insert(0, formatted)
        return True

    def save(self):
        try:
            qty_str = self.quantity_var.get().replace(',', '.')
            price_str = self.unit_price_var.get().replace(',', '.')
            quantity = float(qty_str or 1)
            unit_price = float(price_str or 0)
            total_amount = quantity * unit_price

            diagnosis = self.diagnosis_var.get().strip()

            conn = self.db.get_connection()
            cursor = conn.cursor()
            cursor.execute('''
                           INSERT INTO medical_records
                           (prisoner_id, service_date, diagnosis, doctor_name, service_name,
                            quantity, unit_price, total_amount, service_number, notes)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                           ''', (
                               self.prisoner_id,
                               datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                               diagnosis,
                               self.doctor_name.get().strip(),
                               self.service_name.get().strip(),
                               quantity,
                               unit_price,
                               total_amount,
                               self.service_number.get().strip(),
                               self.notes.get("1.0", tk.END).strip()
                           ))
            conn.commit()
            conn.close()

            messagebox.showinfo("Успех", "Медицинская услуга сохранена")
            self.refresh()
            self.win.destroy()
        except ValueError:
            messagebox.showerror("Ошибка", "Введите корректные числа в полях «Количество» и «Стоимость»")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить:\n{str(e)}")
# ====================== ОКНО ФОРМИРОВАНИЯ ОТЧЁТА ======================
class ReportDialog:
    def __init__(self, parent, main_app):
        self.main_app = main_app
        self.db = main_app.db

        self.win = tk.Toplevel(parent)
        self.win.title("Формирование отчёта")
        self.win.geometry("670x520")
        self.win.grab_set()
        self.win.resizable(True, True)

        ttk.Label(self.win, text="Параметры отчёта", font=("Arial", 14, "bold")).pack(pady=15)

        # Тип отчёта
        ttk.Label(self.win, text="Тип отчёта:").pack(anchor="w", padx=30, pady=(10, 5))
        self.report_combo = ttk.Combobox(self.win, width=55, state="readonly")
        self.report_combo['values'] = [
            "1. Текущие остатки на складе",
            "2. Расходование медикаментов",
            "3. Перемещения за период",
            "4. Учёт медицинской помощи"
        ]
        self.report_combo.current(0)
        self.report_combo.pack(padx=30, fill="x")
        self.report_combo.bind("<<ComboboxSelected>>", self.update_filters)

        # Период (даты)
        date_frame = ttk.LabelFrame(self.win, text="Период", padding=10)
        date_frame.pack(fill="x", padx=30, pady=12)

        ttk.Label(date_frame, text="С:").pack(side=tk.LEFT)
        self.date_from = ttk.Entry(date_frame, width=12)
        self.date_from.pack(side=tk.LEFT, padx=8)
        self.date_from.insert(0, "01.01.2025")
        self.date_from.configure(validate="key", validatecommand=self.main_app.validate_date_cmd)

        ttk.Label(date_frame, text="По:").pack(side=tk.LEFT)
        self.date_to = ttk.Entry(date_frame, width=12)
        self.date_to.pack(side=tk.LEFT, padx=8)
        today = datetime.now().strftime("%d.%m.%Y")
        self.date_to.insert(0, today)
        self.date_to.configure(validate="key", validatecommand=self.main_app.validate_date_cmd)

        # === Фильтры (Медчасть + Заключённый) ===
        filter_frame = ttk.Frame(self.win)
        filter_frame.pack(fill="x", padx=30, pady=8)

        # Медчасть
        self.part_label = ttk.Label(filter_frame, text="Медчасть:")
        self.part_label.pack(anchor="w", pady=(8, 3))
        self.part_var = tk.StringVar()
        self.part_combo = ttk.Combobox(filter_frame, textvariable=self.part_var, width=50, state="readonly")
        self.part_combo.pack(fill="x")

        # Заключённый
        self.prisoner_label = ttk.Label(filter_frame, text="Заключённый:")
        self.prisoner_label.pack(anchor="w", pady=(12, 3))
        self.prisoner_combo = ttk.Combobox(filter_frame, width=50, state="readonly")
        self.prisoner_combo.pack(fill="x")

        # Кнопки в самом низу
        btn_frame = ttk.Frame(self.win)
        btn_frame.pack(pady=30)
        ttk.Button(btn_frame, text="📊 Сформировать отчёт", command=self.generate).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="Отмена", command=self.win.destroy).pack(side=tk.LEFT, padx=10)

        self.load_filters()
        self.update_filters()  # применяем начальное состояние

    def load_filters(self):
        conn = self.db.get_connection()
        cursor = conn.cursor()

        # Медчасти
        cursor.execute("SELECT name FROM med_parts ORDER BY name")
        parts = ["Все медчасти"] + [r["name"] for r in cursor.fetchall()]
        self.part_combo['values'] = parts
        self.part_combo.current(0)

        # Заключённые
        cursor.execute("SELECT id, full_name, prisoner_id FROM prisoners ORDER BY full_name")
        self.prisoners = ["Все заключённые"] + [
            f"{r['full_name']} ({r['prisoner_id']})" for r in cursor.fetchall()
        ]
        self.prisoner_combo['values'] = self.prisoners
        self.prisoner_combo.current(0)

        conn.close()

    def update_filters(self, event=None):
        idx = self.report_combo.current()

        if idx == 3:  # Учёт медицинской помощи
            self.part_label.pack_forget()
            self.part_combo.pack_forget()
            self.prisoner_label.pack(anchor="w", pady=(12, 3))
            self.prisoner_combo.pack(fill="x")
        else:
            self.part_label.pack(anchor="w", pady=(8, 3))
            self.part_combo.pack(fill="x")
            self.prisoner_label.pack_forget()
            self.prisoner_combo.pack_forget()

    def generate(self):
        idx = self.report_combo.current()
        d_from = self.date_from.get().strip()
        d_to = self.date_to.get().strip()
        part = self.part_var.get()

        d_from_sql = self.main_app.convert_date_for_query(d_from)
        d_to_sql = self.main_app.convert_date_for_query(d_to)

        if idx == 0:
            self.main_app.report_stock()
        elif idx == 1:
            self.main_app.report_expenditures(d_from_sql, d_to_sql, part)
        elif idx == 2:
            self.main_app.report_movements(d_from_sql, d_to_sql, part)
        elif idx == 3:
            prisoner_filter = self.prisoner_combo.get()
            self.main_app.report_medical_records(d_from_sql, d_to_sql, prisoner_filter)

        self.win.destroy()

if __name__ == "__main__":
    app = MainWindow()
    app.run()


